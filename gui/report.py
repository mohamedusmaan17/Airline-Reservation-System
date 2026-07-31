import os
import sys

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

from database.db import connect_db


class ReportWindow:

    def __init__(self, root):

        self.root = root

        self.root.title("Reports")

        self.root.geometry("1200x650")

        self.root.configure(bg="white")

        # ==========================
        # Header
        # ==========================

        header = tk.Label(
            root,
            text="REPORT MANAGEMENT",
            bg="#003366",
            fg="white",
            font=("Arial", 24, "bold"),
            pady=15
        )

        header.pack(fill="x")

        # ==========================
        # Buttons
        # ==========================

        button_frame = tk.Frame(root, bg="white")

        button_frame.pack(pady=20)

        tk.Button(
            button_frame,
            text="Passenger Report",
            width=18,
            command=self.passenger_report
        ).grid(row=0, column=0, padx=10)

        tk.Button(
            button_frame,
            text="Flight Report",
            width=18,
            command=self.flight_report
        ).grid(row=0, column=1, padx=10)

        tk.Button(
            button_frame,
            text="Booking Report",
            width=18,
            command=self.booking_report
        ).grid(row=0, column=2, padx=10)

        tk.Button(
            button_frame,
            text="Payment Report",
            width=18,
            command=self.payment_report
        ).grid(row=0, column=3, padx=10)

        tk.Button(
            button_frame,
            text="Export Excel",
            width=18,
            bg="green",
            fg="white",
            command=self.export_excel
        ).grid(row=0, column=4, padx=10)

        tk.Button(
            button_frame,
            text="Export PDF",
            width=18,
            bg="#d9534f",
            fg="white",
            command=self.export_pdf
        ).grid(row=0, column=5, padx=10)

        # ==========================
        # Treeview
        # ==========================

        table_frame = tk.Frame(root)

        table_frame.pack(fill="both", expand=True, padx=20, pady=20)

        scroll_x = tk.Scrollbar(table_frame, orient="horizontal")

        scroll_y = tk.Scrollbar(table_frame)

        self.report_table = ttk.Treeview(
            table_frame,
            xscrollcommand=scroll_x.set,
            yscrollcommand=scroll_y.set
        )

        scroll_x.pack(side="bottom", fill="x")

        scroll_y.pack(side="right", fill="y")

        scroll_x.config(command=self.report_table.xview)

        scroll_y.config(command=self.report_table.yview)

        self.report_table.configure(
            xscrollcommand=scroll_x.set,
            yscrollcommand=scroll_y.set
        )
        self.report_table.pack(fill="both", expand=True)

        scroll_x.pack(side="bottom", fill="x")
        scroll_y.pack(side="right", fill="y")
     # ==========================
     #         # Empty Methods
    # ==========================

    def clear_table(self):
        self.report_table.delete(*self.report_table.get_children())

    def passenger_report(self):

        try:

            conn = connect_db()
            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    passenger_id,
                    first_name,
                    last_name,
                    gender,
                    date_of_birth,
                    phone,
                    email,
                    passport_number,
                    nationality
                FROM passengers
            """)

            rows = cursor.fetchall()

            self.report_table.delete(*self.report_table.get_children())

            self.report_table["columns"] = (
                "ID",
                "First Name",
                "Last Name",
                "Gender",
                "DOB",
                "Phone",
                "Email",
                "Passport",
                "Nationality"
            )

            self.report_table["show"] = "headings"

            widths = [70,120,120,90,100,120,200,120,120]

            for col, width in zip(self.report_table["columns"], widths, strict=False):
                self.report_table.heading(col, text=col)
                self.report_table.column(
                    col,
                    width=width,
                    anchor="center"
                )
            for row in rows:
                self.report_table.insert("", "end", values=row)

            conn.close()

        except Exception as e:
            messagebox.showerror("Database Error", str(e))

    def flight_report(self):

        try:

            from database.db import connect_db

            conn = connect_db()

            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    f.flight_id,
                    a.airline_name,
                    ap1.airport_name,
                    ap2.airport_name,
                    f.flight_number,
                    f.departure_time,
                    f.arrival_time,
                    f.total_seats,
                    f.available_seats,
                    f.ticket_price
                FROM flights f
                JOIN airlines a
                    ON f.airline_id = a.airline_id
                JOIN airports ap1
                    ON f.source_airport = ap1.airport_id
                JOIN airports ap2
                    ON f.destination_airport = ap2.airport_id
            """)

            rows = cursor.fetchall()

            self.report_table.delete(*self.report_table.get_children())

            columns = (
                "ID",
                "Airline",
                "Source",
                "Destination",
                "Flight No",
                "Departure",
                "Arrival",
                "Total Seats",
                "Available",
                "Price"
            )

            self.report_table["columns"] = columns
            self.report_table["show"] = "headings"

            for col in columns:
                self.report_table.heading(col, text=col)
                self.report_table.column(col, width=120)

            for row in rows:
                self.report_table.insert("", tk.END, values=row)

            conn.close()

        except Exception as e:

            import traceback
            traceback.print_exc()

            messagebox.showerror(
                "Database Error",
                str(e)
            )
    def booking_report(self):

        try:

            from database.db import connect_db

            conn = connect_db()

            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    b.booking_id,
                    CONCAT(p.first_name,' ',p.last_name),
                    f.flight_number,
                    b.seat_number,
                    b.booking_date,
                    b.booking_status
                FROM bookings b
                JOIN passengers p
                    ON b.passenger_id = p.passenger_id
                JOIN flights f
                    ON b.flight_id = f.flight_id
                ORDER BY b.booking_id
            """)

            rows = cursor.fetchall()

            self.report_table.delete(*self.report_table.get_children())

            columns = (
                "Booking ID",
                "Passenger",
                "Flight",
                "Seat",
                "Booking Date",
                "Status"
            )

            self.report_table["columns"] = columns
            self.report_table["show"] = "headings"

            widths = [90, 180, 100, 80, 170, 120]

            for col, width in zip(columns, widths, strict=False):
                self.report_table.heading(col, text=col)
                self.report_table.column(col, width=width, anchor="center")

            for row in rows:
                self.report_table.insert("", tk.END, values=row)

            conn.close()

        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Database Error", str(e))
    def payment_report(self):

        try:

            from database.db import connect_db

            conn = connect_db()

            cursor = conn.cursor()

            cursor.execute("""
                SELECT
                    pay.payment_id,
                    CONCAT(p.first_name,' ',p.last_name) AS Passenger,
                    f.flight_number,
                    pay.amount,
                    pay.payment_method,
                    pay.payment_date,
                    pay.payment_status
                FROM payments pay

                JOIN bookings b
                    ON pay.booking_id = b.booking_id

                JOIN passengers p
                    ON b.passenger_id = p.passenger_id

                JOIN flights f
                    ON b.flight_id = f.flight_id

                ORDER BY pay.payment_id
            """)

            rows = cursor.fetchall()

            self.report_table.delete(*self.report_table.get_children())

            columns = (
                "Payment ID",
                "Passenger",
                "Flight",
                "Amount",
                "Method",
                "Payment Date",
                "Status"
            )

            self.report_table["columns"] = columns
            self.report_table["show"] = "headings"

            widths = [90, 180, 100, 100, 140, 180, 120]

            for col, width in zip(columns, widths, strict=False):
                self.report_table.heading(col, text=col)
                self.report_table.column(
                    col,
                    width=width,
                    anchor="center",
                    stretch=True
                )
            for row in rows:
                self.report_table.insert("", tk.END, values=row)

            conn.close()

        except Exception as e:

            import traceback
            traceback.print_exc()

            messagebox.showerror(
                "Database Error",
                str(e)
            )

    def export_excel(self):

        if not self.report_table.get_children():
            messagebox.showwarning(
                "Warning",
                "No data available to export."
            )
            return

        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )

        if not file:
            return

        workbook = Workbook()
        sheet = workbook.active
        from openpyxl.styles import Font

        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.title = "Report"

        # Write column headings
        columns = self.report_table["columns"]

        for col_num, heading in enumerate(columns, start=1):
            sheet.cell(row=1, column=col_num).value = heading

        # Write rows
        for row_num, item in enumerate(self.report_table.get_children(), start=2):
            values = self.report_table.item(item)["values"]

            for col_num, value in enumerate(values, start=1):
                sheet.cell(row=row_num, column=col_num).value = value

        workbook.save(file)

        messagebox.showinfo(
            "Success",
            "Excel Report Exported Successfully!"
        )
    def export_pdf(self):

        rows = []

        columns = self.report_table["columns"]

        rows.append(columns)

        for item in self.report_table.get_children():
            rows.append(self.report_table.item(item)["values"])

        if len(rows) <= 1:
            messagebox.showerror(
                "Error",
                "No data available to export."
            )
            return

        file = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")]
        )

        if not file:
            return

        doc = SimpleDocTemplate(file)

        elements = []

        styles = getSampleStyleSheet()

        title = Paragraph(
            "<b><font size=18>AIRLINE RESERVATION SYSTEM</font></b>",
            styles["Title"]
        )

        elements.append(title)

        elements.append(Paragraph("<br/><br/>", styles["Normal"]))

        table = Table(rows)

        table.setStyle(TableStyle([

            ("BACKGROUND", (0,0), (-1,0), colors.darkblue),

            ("TEXTCOLOR", (0,0), (-1,0), colors.white),

            ("ALIGN", (0,0), (-1,-1), "CENTER"),

            ("GRID", (0,0), (-1,-1), 1, colors.black),

            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0,0), (-1,0), 8),

            ("BACKGROUND", (0,1), (-1,-1), colors.beige)

        ]))

        elements.append(table)

        doc.build(elements)

        messagebox.showinfo(
            "Success",
            "PDF exported successfully!"
        )
